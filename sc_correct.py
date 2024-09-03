import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
from openpyxl.drawing.image import Image
import sys
import win32com.client as win32

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

def process_sc_correct():
    # Browse and select files
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if not files:
        return

    file_listbox.delete(0, tk.END)
    for file in files:
        file_listbox.insert(tk.END, file)

    # Save files
    save_directory = filedialog.askdirectory()
    if not save_directory:
        return

    target_directory = os.path.join(save_directory, 'sc-corrected')
    os.makedirs(target_directory, exist_ok=True)

    files = file_listbox.get(0, tk.END)
    for file_path in files:
        input_workbook = openpyxl.load_workbook(file_path)
        sheet = input_workbook["SC"]

        find_text1 = '1134 Budapest, Váci út 33., Hungary'
        find_text2 = 'TT:15% LC:85%'
        find_text3 = 'Seller：SAIC Motor Central and Eastern Europe Kft.'
        find_text4 = 'Buyer:'
        image_path = resource_path("image.png")

        replace_text1 = '1138 Budapest, Esztergomi út 31-39. Hungary'
        replace_text2 = 'TT:5% LC:95%'

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value).strip()

                    if find_text2 in cell_value:
                        cell.value = replace_text2

                    if cell_value == find_text1:
                        cell.value = replace_text1

                    if find_text3 in cell_value:
                        image = Image(image_path)
                        image_cell_col = cell.column + 5
                        image_cell_row = cell.row
                        image.anchor = sheet.cell(row=image_cell_row, column=image_cell_col).coordinate
                        sheet.add_image(image)

                    if find_text4 in cell_value:
                        buyer_cell = sheet.cell(row=cell.row, column=cell.column + 3)
                        buyer_text = str(buyer_cell.value).strip()

                        if len(buyer_text) > 20:
                            mid_index = len(buyer_text) // 2
                            if buyer_text[mid_index] != ' ':
                                mid_index = buyer_text.rfind(' ', 0, mid_index)

                            part1 = buyer_text[:mid_index].strip()
                            part2 = buyer_text[mid_index:].strip()

                            buyer_cell.value = f"{part1}\n{part2}"
                            buyer_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                        else:
                            buyer_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

                        # Resize row and column to fit text
                        max_length = 0
                        for row in sheet.iter_rows():
                            cell = row[buyer_cell.column - 1]  # OpenPyXL is zero-indexed for column access
                            if cell.value:
                                try:
                                    max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                        adjusted_width = max_length + 2  # Add a bit of padding
                        sheet.column_dimensions[openpyxl.utils.get_column_letter(buyer_cell.column)].width = adjusted_width

                        # Adjust row height
                        max_height = 15  # Default height; adjust if needed
                        lines = str(buyer_cell.value).split('\n')
                        num_lines = len(lines)
                        max_height = max(max_height, num_lines * 15)  # Adjust line height as needed
                        sheet.row_dimensions[buyer_cell.row].height = max_height

        base_name = os.path.basename(file_path)
        new_path = os.path.join(target_directory, base_name)
        input_workbook.save(new_path)

        # Convert the saved Excel file to PDF
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        try:
            workbook = excel.Workbooks.Open(new_path)
            pdf_path = os.path.splitext(new_path)[0] + '.pdf'
            pdf_path = os.path.join(target_directory, os.path.basename(pdf_path))
            workbook.ExportAsFixedFormat(0, pdf_path)  # 0 stands for PDF format
        finally:
            workbook.Close(SaveChanges=False)
            excel.Application.Quit()

    messagebox.showinfo("Success", "Files have been saved and converted to PDF successfully.")

root = tk.Tk()
root.title("SC Correct")
root.geometry("400x400")

frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

scrollbar.config(command=file_listbox.yview)

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Process Files", command=process_sc_correct)
select_files_button.pack(side=tk.LEFT, padx=10)

root.mainloop()
