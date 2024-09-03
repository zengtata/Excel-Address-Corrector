import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import sys

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)

def save_files():
    save_directory = filedialog.askdirectory()
    if save_directory:
        target_directory = os.path.join(save_directory, 'ci-corrected')
        os.makedirs(target_directory, exist_ok=True)

        files = file_listbox.get(0, tk.END)
        for file_path in files:
            input_workbook = load_workbook(file_path)
            sheet = input_workbook["CI"]

            # Old and new addresses
            old_address = "1134 Budapest, Váci út 33. HUNGARY"
            new_address = "1138 Budapest, Esztergomi út 31-39. Hungary"

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Replace old address with new address in the cell's value
                        if old_address in cell.value:
                            cell.value = cell.value.replace(old_address, new_address)

            base_name = os.path.basename(file_path)
            new_path = os.path.join(target_directory, base_name)
            input_workbook.save(new_path)

        tk.messagebox.showinfo("Success", "Files have been saved successfully.")

root = tk.Tk()
root.title("CIPL Correct")
root.geometry("400x400")

frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

scrollbar.config(command=file_listbox.yview)

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_files_button = tk.Button(button_frame, text="Save As", command=save_files)
save_files_button.pack(side=tk.LEFT, padx=10)

root.mainloop()
