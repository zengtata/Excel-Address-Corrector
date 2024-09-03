import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import sys

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

def browse_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        file_listbox.delete(0, tk.END)
        for file in files:
            file_listbox.insert(tk.END, file)

def save_files():
    save_directory = filedialog.askdirectory()
    if save_directory:
        target_directory = os.path.join(save_directory, 'pi-corrected')
        os.makedirs(target_directory, exist_ok=True)

        files = file_listbox.get(0, tk.END)
        for file_path in files:
            input_workbook = load_workbook(file_path)
            sheet = input_workbook["PI"]

            find_text1 = '1134 Budapest, Váci út 33., Hungary'
            find_text2 = '15% by T/T'
            find_text3 = '85% by L/C'
            find_text4 = 'Bank Information:'
            image_path = resource_path("image.png")

            replace_text1 = '1138 Budapest, Esztergomi út 31-39. Hungary'
            replace_text2 = '5% by T/T'
            replace_text3 = '95% by L/C'

            total_payment = 0
            tt = 0
            lc = 0

            no = []
            unit_price = []
            price = []
            qty = []

            for col in sheet.iter_cols(values_only=True):
                for idx, cell in enumerate(col):
                    if cell == "No.":
                        for next_idx in range(idx + 1, len(col)):
                            if col[next_idx] is not None:
                                no.append(col[next_idx])
                            if next_idx + 1 < len(col) and col[next_idx + 1] == "TOTAL Qty:":
                                break
                    elif cell == "Unit Price":
                        for next_idx in range(idx + 1, len(col)):
                            if col[next_idx] is not None:
                                unit_price.append(col[next_idx])
                            else:
                                break
                    elif cell == "Qty":
                        for next_idx in range(idx + 1, idx + len(no) + 1):
                            if next_idx < len(col) and col[next_idx] is not None:
                                qty.append(col[next_idx])
                            else:
                                break

            for i in range(len(unit_price)):
                price.append(unit_price[i] * qty[i])

            for i in range(len(price)):
                total_payment += price[i]

            tt = 0.05 * total_payment
            lc = 0.95 * total_payment

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip()
                        if find_text2 in cell_value:
                            cell_to_update = sheet.cell(row=cell.row, column=cell.column + 3)
                            cell_to_update.value = tt
                            cell.value = replace_text2

                        if find_text3 in cell_value:
                            cell_to_update = sheet.cell(row=cell.row, column=cell.column + 3)
                            cell_to_update.value = lc
                            cell.value = replace_text3

                        if cell_value == find_text1:
                            cell.value = replace_text1

                        if find_text4 in cell_value:
                            image = Image(image_path)
                            image_cell_col = cell.column + 4
                            image_cell_row = cell.row
                            image.anchor = sheet.cell(row=image_cell_row, column=image_cell_col).coordinate
                            sheet.add_image(image)

            base_name = os.path.basename(file_path)
            new_path = os.path.join(target_directory, base_name)
            input_workbook.save(new_path)

        tk.messagebox.showinfo("Success", "Files have been saved successfully.")


root = tk.Tk()
root.title("PI Correct")
root.geometry("400x400")

frame = tk.Frame(root)
frame.pack(pady=10, padx=10, fill='both', expand=True)

scrollbar = tk.Scrollbar(frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

file_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set)
file_listbox.pack(side=tk.LEFT, fill='both', expand=True)

scrollbar.config(command=file_listbox.yview)

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

select_files_button = tk.Button(button_frame, text="Select Files", command=browse_files)
select_files_button.pack(side=tk.LEFT, padx=10)

save_files_button = tk.Button(button_frame, text="Save As", command=save_files)
save_files_button.pack(side=tk.LEFT, padx=10)

root.mainloop()
